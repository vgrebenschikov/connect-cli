# -*- coding: utf-8 -*-

# This file is part of the Ingram Micro Cloud Blue Connect connect-cli.
# Copyright (c) 2019-2022 Ingram Micro. All Rights Reserved.

from collections import namedtuple

import click
from tqdm import tqdm

from connect.client import ClientError

from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from connect.cli.core.constants import DEFAULT_BAR_FORMAT
from connect.cli.plugins.shared.sync_stats import SynchronizerStats
from connect.cli.plugins.shared.exceptions import SheetNotFoundError
from connect.cli.plugins.shared.constants import ATTRIBUTES_SHEET_COLUMNS


class TranslationAttributesSynchronizer:
    """
    Synchronize the attributes of a translation from excel file.
    """
    def __init__(self, client, silent, stats=None):
        self._client = client
        self._silent = silent
        self._wb = None
        self._ws = None
        if stats is None:
            stats = SynchronizerStats()
        self._mstats = stats['Attributes']

    def open(self, input_file, worksheet):
        self._open_workbook(input_file)
        if worksheet not in self._wb.sheetnames:
            raise SheetNotFoundError(f"File does not contain worksheet '{worksheet}' to synchronize, skipping")
        self._ws = self._wb[worksheet]
        self._validate_attributes_worksheet(self._ws)

    def save(self, output_file):
        self._wb.save(output_file)

    def sync(self, translation_id):
        attributes = self._collect_attributes_to_update(self._ws)
        if attributes:
            self._update_attributes(translation_id, attributes, self._ws)

    def _open_workbook(self, input_file):
        try:
            self._wb = load_workbook(input_file, data_only=True)
        except InvalidFileException as ife:
            raise click.ClickException(str(ife))
        except BadZipFile:
            raise click.ClickException(f'{input_file} is not a valid xlsx file.')

    @staticmethod
    def _validate_attributes_worksheet(ws):
        for col_idx, header in enumerate(ATTRIBUTES_SHEET_COLUMNS, 1):
            if header == 'original value':
                continue
            cell = ws.cell(1, col_idx)
            if cell.value != header:
                raise click.ClickException(
                    f"Column '{cell.coordinate}' must be '{header}', but it is '{cell.value}'",
                )

    def _collect_attributes_to_update(self, ws):
        AttributeRow = namedtuple(
            'AttributeRow',
            (header.replace(' ', '_').lower() for header in ATTRIBUTES_SHEET_COLUMNS),
        )

        progress = tqdm(
            enumerate(ws.iter_rows(min_row=2, values_only=True), 2),
            total=ws.max_row - 1, disable=self._silent, leave=True, bar_format=DEFAULT_BAR_FORMAT,
        )

        attributes = {}
        for row_idx, row in progress:
            row = AttributeRow(*row)
            progress.set_description(f'Process attribute {row.key}')
            if row.action == 'update':
                attributes[row_idx] = {'key': row.key, 'value': row.value, 'comment': row.comment}
            else:
                self._mstats.skipped()

        return attributes

    def _update_attributes(self, translation_id, attributes, ws):
        try:
            translation_res = self._client.ns('localization').translations[translation_id]
            translation_res.attributes.bulk_update(list(attributes.values()))
            self._mstats.updated(len(attributes))
            for row_idx in attributes.keys():
                self._update_attributes_sheet_row(ws, row_idx)
        except ClientError as e:
            self._mstats.error(
                f'Error while updating attributes: {str(e)}',
                range(1, len(attributes) + 1),
            )

    @staticmethod
    def _update_attributes_sheet_row(ws, row_idx):
        ws.cell(row_idx, 3, value='-')